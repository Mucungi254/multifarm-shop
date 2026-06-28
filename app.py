from flask import Flask, render_template, request, redirect, url_for, flash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from config import Config
from datetime import datetime, date, timedelta
import openpyxl
from sqlalchemy import text

app = Flask(__name__)
app.config.from_object(Config)

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in.'

# ---------- User Class ----------
class User(UserMixin):
    def __init__(self, id, username, role):
        self.id = id
        self.username = username
        self.role = role

@login_manager.user_loader
def load_user(user_id):
    result = db.session.execute(text("SELECT id, username, role FROM users WHERE id = :id"), {'id': user_id})
    user = result.fetchone()
    if user:
        return User(user.id, user.username, user.role)
    return None

# ---------- Ensure users exist ----------
def init_users():
    # Check if users exist
    result = db.session.execute(text("SELECT id FROM users WHERE username = 'admin'"))
    if not result.fetchone():
        db.session.execute(text(
            "INSERT INTO users (username, password_hash, role, full_name) VALUES (:username, :hash, :role, :full_name)"
        ), {'username': 'admin', 'hash': generate_password_hash('admin123'), 'role': 'admin', 'full_name': 'Mucungi'})
    result = db.session.execute(text("SELECT id FROM users WHERE username = 'sarah'"))
    if not result.fetchone():
        db.session.execute(text(
            "INSERT INTO users (username, password_hash, role, full_name) VALUES (:username, :hash, :role, :full_name)"
        ), {'username': 'sarah', 'hash': generate_password_hash('sarah123'), 'role': 'agent', 'full_name': 'Sarah'})
    result = db.session.execute(text("SELECT id FROM users WHERE username = 'muchiri'"))
    if not result.fetchone():
        db.session.execute(text(
            "INSERT INTO users (username, password_hash, role, full_name) VALUES (:username, :hash, :role, :full_name)"
        ), {'username': 'muchiri', 'hash': generate_password_hash('muchiri123'), 'role': 'secretary', 'full_name': 'Muchiri'})
    db.session.commit()

# ---------- Routes ----------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        result = db.session.execute(text("SELECT id, username, password_hash, role FROM users WHERE username = :username"), {'username': username})
        user = result.fetchone()
        if user and check_password_hash(user.password_hash, password):
            login_user(User(user.id, user.username, user.role))
            flash('Logged in successfully.', 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid credentials.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/', methods=['GET', 'POST'])
@login_required
def dashboard():
    date_str = request.args.get('date', date.today().isoformat())
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except:
        selected_date = date.today()

    # ---- POST: Save Record ----
    if request.method == 'POST':
        if current_user.role not in ['admin', 'agent']:
            flash('Not allowed.', 'danger')
            return redirect(url_for('dashboard', date=selected_date))

        john_am = float(request.form.get('john_am_litres', 0) or 0)
        john_pm = float(request.form.get('john_pm_litres', 0) or 0)
        sales_am_litres = float(request.form.get('sales_am_litres', 0) or 0)
        sales_pm_litres = float(request.form.get('sales_pm_litres', 0) or 0)

        sales_am_amount = sales_am_litres * 60
        sales_pm_amount = sales_pm_litres * 60

        otc_paybill = float(request.form.get('otc_paybill', 0) or 0)
        sarah_cash = float(request.form.get('sarah_cash_deposit', 0) or 0)
        bank_stmt = float(request.form.get('bank_statement', 0) or 0)

        # Check if record exists
        result = db.session.execute(text("SELECT id FROM daily_records WHERE record_date = :date"), {'date': selected_date})
        existing = result.fetchone()

        if existing:
            db.session.execute(text("""
                UPDATE daily_records SET
                    john_am_litres = :jam, john_pm_litres = :jpm,
                    sales_am_litres = :sam, sales_am_amount = :saa,
                    sales_pm_litres = :spm, sales_pm_amount = :spa,
                    otc_paybill = :otc, sarah_cash_deposit = :cash,
                    bank_statement = :bank
                WHERE record_date = :date
            """), {
                'jam': john_am, 'jpm': john_pm,
                'sam': sales_am_litres, 'saa': sales_am_amount,
                'spm': sales_pm_litres, 'spa': sales_pm_amount,
                'otc': otc_paybill, 'cash': sarah_cash, 'bank': bank_stmt,
                'date': selected_date
            })
        else:
            db.session.execute(text("""
                INSERT INTO daily_records (
                    record_date, john_am_litres, john_pm_litres,
                    sales_am_litres, sales_am_amount,
                    sales_pm_litres, sales_pm_amount,
                    otc_paybill, sarah_cash_deposit, bank_statement
                ) VALUES (
                    :date, :jam, :jpm, :sam, :saa, :spm, :spa, :otc, :cash, :bank
                )
            """), {
                'date': selected_date,
                'jam': john_am, 'jpm': john_pm,
                'sam': sales_am_litres, 'saa': sales_am_amount,
                'spm': sales_pm_litres, 'spa': sales_pm_amount,
                'otc': otc_paybill, 'cash': sarah_cash, 'bank': bank_stmt
            })
        db.session.commit()
        flash('Record saved.', 'success')
        return redirect(url_for('dashboard', date=selected_date.isoformat()))

    # ---- GET: Load Record ----
    record = db.session.execute(text("SELECT * FROM daily_records WHERE record_date = :date"), {'date': selected_date}).fetchone()

    # Farmer deliveries
    deliveries = db.session.execute(text("""
        SELECT fd.*, f.member_no, f.name
        FROM farmer_deliveries fd
        JOIN farmers f ON fd.farmer_id = f.id
        WHERE fd.record_date = :date
        ORDER BY f.member_no
    """), {'date': selected_date}).fetchall()

    farmers = db.session.execute(text("SELECT id, member_no, name FROM farmers ORDER BY member_no")).fetchall()

    # ---- Calculate totals ----
    if record:
        farmer_total = sum(d.litres for d in deliveries)
        john_total = (record.john_am_litres or 0) + (record.john_pm_litres or 0)
        total_milk_in = john_total + farmer_total
        total_sold = (record.sales_am_litres or 0) + (record.sales_pm_litres or 0)
        total_revenue = (record.sales_am_amount or 0) + (record.sales_pm_amount or 0)
        reported_deposits = (record.otc_paybill or 0) + (record.sarah_cash_deposit or 0)
        money_not_deposited = total_revenue - reported_deposits
        milk_left = total_milk_in - total_sold
        bank_stmt = record.bank_statement or 0
        deposit_discrepancy = bank_stmt - reported_deposits if bank_stmt else None
    else:
        farmer_total = john_total = total_milk_in = total_sold = total_revenue = reported_deposits = money_not_deposited = milk_left = bank_stmt = 0
        deposit_discrepancy = None

    return render_template('dashboard.html',
                           date=selected_date,
                           record=record,
                           deliveries=deliveries,
                           farmers=farmers,
                           farmer_total=farmer_total,
                           john_total=john_total,
                           total_milk_in=total_milk_in,
                           total_sold=total_sold,
                           total_revenue=total_revenue,
                           reported_deposits=reported_deposits,
                           money_not_deposited=money_not_deposited,
                           milk_left=milk_left,
                           bank_stmt=bank_stmt,
                           deposit_discrepancy=deposit_discrepancy,
                           timedelta=timedelta)

# ---------- Add Farmer Delivery ----------
@app.route('/add_farmer_delivery', methods=['POST'])
@login_required
def add_farmer_delivery():
    if current_user.role not in ['admin', 'agent']:
        flash('Not allowed.', 'danger')
        return redirect(url_for('dashboard'))

    date_str = request.form.get('date')
    farmer_id = request.form.get('farmer_id')
    litres = float(request.form.get('litres', 0) or 0)

    if not date_str or not farmer_id or litres <= 0:
        flash('Please fill all fields.', 'warning')
        return redirect(url_for('dashboard', date=date_str))

    # Ensure daily record exists
    result = db.session.execute(text("SELECT id FROM daily_records WHERE record_date = :date"), {'date': date_str})
    if not result.fetchone():
        db.session.execute(text("INSERT INTO daily_records (record_date) VALUES (:date)"), {'date': date_str})
        db.session.commit()
        flash('Note: A blank daily record was created for this date.', 'info')

    db.session.execute(text("INSERT INTO farmer_deliveries (record_date, farmer_id, litres) VALUES (:date, :fid, :litres)"),
                       {'date': date_str, 'fid': farmer_id, 'litres': litres})
    db.session.commit()
    flash('Farmer delivery added.', 'success')
    return redirect(url_for('dashboard', date=date_str))

# ---------- Delete Farmer Delivery ----------
@app.route('/delete_farmer_delivery/<int:delivery_id>')
@login_required
def delete_farmer_delivery(delivery_id):
    if current_user.role not in ['admin', 'agent']:
        flash('Not allowed.', 'danger')
        return redirect(url_for('dashboard'))

    result = db.session.execute(text("SELECT record_date FROM farmer_deliveries WHERE id = :id"), {'id': delivery_id})
    row = result.fetchone()
    if row:
        db.session.execute(text("DELETE FROM farmer_deliveries WHERE id = :id"), {'id': delivery_id})
        db.session.commit()
        flash('Delivery removed.', 'info')
        date_str = row.record_date.isoformat()
    else:
        date_str = date.today().isoformat()
    return redirect(url_for('dashboard', date=date_str))

# ---------- Manage Farmers ----------
@app.route('/farmers', methods=['GET', 'POST'])
@login_required
def manage_farmers():
    if current_user.role != 'admin':
        flash('Admin access only.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST' and 'add_farmer' in request.form:
        member_no = request.form.get('member_no')
        name = request.form.get('name')
        rider = request.form.get('rider', 'B')
        if member_no and name:
            db.session.execute(text("INSERT INTO farmers (member_no, name, rider) VALUES (:no, :name, :rider)"),
                               {'no': member_no, 'name': name, 'rider': rider})
            db.session.commit()
            flash('Farmer added manually.', 'success')
        else:
            flash('Member No and Name are required.', 'warning')
        return redirect(url_for('manage_farmers'))

    farmers = db.session.execute(text("SELECT * FROM farmers ORDER BY member_no")).fetchall()
    return render_template('farmers_list.html', farmers=farmers)

# ---------- Import Farmers ----------
@app.route('/import_farmers', methods=['GET', 'POST'])
@login_required
def import_farmers():
    if current_user.role != 'admin':
        flash('Admin access only.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('Please select a file.', 'warning')
            return redirect(url_for('import_farmers'))

        if not file.filename.lower().endswith('.xlsx'):
            flash('Please upload an Excel (.xlsx) file.', 'danger')
            return redirect(url_for('import_farmers'))

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            flash(f'Error reading file: {e}', 'danger')
            return redirect(url_for('import_farmers'))

        # Find header row
        header_row = None
        header_row_num = None
        for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            if row and any(cell and ('member' in str(cell).lower() or 'name' in str(cell).lower()) for cell in row):
                header_row = row
                header_row_num = row_num
                break

        if not header_row:
            flash('Could not find header row with "member_no" and "name".', 'danger')
            return redirect(url_for('import_farmers'))

        col_map = {}
        for idx, col in enumerate(header_row):
            col_lower = str(col).lower().strip()
            if col_lower in ['member_no', 'memberno', 'member no', 'member number']:
                col_map['member_no'] = idx
            elif col_lower in ['name', 'fullname', 'farmer name']:
                col_map['name'] = idx
            elif col_lower in ['rider']:
                col_map['rider'] = idx

        if 'member_no' not in col_map or 'name' not in col_map:
            flash('Excel must have "member_no" and "name" columns.', 'danger')
            return redirect(url_for('import_farmers'))

        added = 0
        skipped = 0

        for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
            member_no = str(row[col_map['member_no']]).strip() if row[col_map['member_no']] else ''
            name = str(row[col_map['name']]).strip() if row[col_map['name']] else ''
            rider = str(row[col_map.get('rider')]).strip() if (col_map.get('rider') is not None and row[col_map['rider']]) else 'B'
            if not member_no:
                continue
            # Check if exists
            result = db.session.execute(text("SELECT id FROM farmers WHERE member_no = :no"), {'no': member_no})
            if result.fetchone():
                skipped += 1
                continue
            db.session.execute(text("INSERT INTO farmers (member_no, name, rider) VALUES (:no, :name, :rider)"),
                               {'no': member_no, 'name': name, 'rider': rider})
            added += 1

        db.session.commit()
        flash(f'Imported {added} farmers. Skipped {skipped} duplicates.', 'success')
        return redirect(url_for('manage_farmers'))

    return render_template('import_farmers.html')

# ---------- Reports ----------
@app.route('/reports')
@login_required
def reports():
    month_str = request.args.get('month', date.today().strftime('%Y-%m'))
    try:
        year, month = map(int, month_str.split('-'))
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year+1, 1, 1)
        else:
            end_date = date(year, month+1, 1)
    except:
        flash('Invalid month format.', 'danger')
        return redirect(url_for('reports'))

    records = db.session.execute(text("""
        SELECT * FROM daily_records
        WHERE record_date >= :start AND record_date < :end
        ORDER BY record_date DESC
    """), {'start': start_date, 'end': end_date}).fetchall()

    processed = []
    for rec in records:
        farmer_sum = db.session.execute(text("SELECT SUM(litres) as total FROM farmer_deliveries WHERE record_date = :date"),
                                        {'date': rec.record_date}).fetchone().total or 0
        john_total = (rec.john_am_litres or 0) + (rec.john_pm_litres or 0)
        total_milk_in = john_total + farmer_sum
        total_sold = (rec.sales_am_litres or 0) + (rec.sales_pm_litres or 0)
        revenue = (rec.sales_am_amount or 0) + (rec.sales_pm_amount or 0)
        reported_deposits = (rec.otc_paybill or 0) + (rec.sarah_cash_deposit or 0)
        milk_left = total_milk_in - total_sold
        processed.append({
            'date': rec.record_date,
            'milk_left': milk_left,
            'total_sold': total_sold,
            'revenue': revenue,
            'reported_deposits': reported_deposits,
            'bank_statement': rec.bank_statement or 0,
            'not_deposited': revenue - reported_deposits,
            'discrepancy': (rec.bank_statement or 0) - reported_deposits if rec.bank_statement else None
        })

    return render_template('reports.html', records=processed, month=month_str)

# ---------- Init ----------
with app.app_context():
    # Ensure tables exist (auto-create)
    db.create_all()
    init_users()

if __name__ == '__main__':
    app.run(debug=True)